from flask import Flask, render_template, request, send_file, url_for
import sys
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime, timedelta
import random
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (NoSuchElementException,
                                        TimeoutException,
                                        WebDriverException)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side

app = Flask(__name__)

def random_delay(min_sec=1, max_sec=2):
    """Random delay between actions"""
    time.sleep(random.uniform(min_sec, max_sec))


def human_like_interaction(driver):
    """Simulate human mouse movements and pauses"""
    try:
        action = ActionChains(driver)
        for _ in range(random.randint(1, 5)):
            x = random.randint(-50, 50)
            y = random.randint(-50, 50)
            action.move_by_offset(x, y).perform()
            time.sleep(random.uniform(0.1, 0.3))

        if random.random() > 0.3:
            scroll_amount = random.randint(200, 600)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            time.sleep(random.uniform(0.5, 1.5))
    except Exception as e:
        print(f"Interaction simulation failed: {str(e)}")


def setup_driver():
    """Configure Chrome with maximum stealth settings"""
    options = uc.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")

    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
    ]
    options.add_argument(f"user-agent={random.choice(user_agents)}")
    options.add_argument(f"--window-size={random.randint(1000, 1400)},{random.randint(800, 1200)}")

    try:
        driver = uc.Chrome(
            options=options,
            headless=False,
            use_subprocess=True
        )

        # Remove webdriver flag and add fake plugins
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                window.navigator.chrome = {
                    runtime: {},
                };
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3]
                });
            """
        })
        return driver
    except Exception as e:
        print(f"❌ Failed to initialize WebDriver: {str(e)}")
        sys.exit(1)


def handle_possible_blocking(driver, current_url):
    """Enhanced blocking handler with cookie clearing and retry"""
    blocking_indicators = [
        "//div[contains(text(), 'Access Denied')]",
        "//div[contains(text(), 'Checking your browser')]",
        "//div[contains(text(), 'Please verify you are a human')]",
        "//iframe[contains(@title, 'recaptcha')]",
        "//div[contains(@class, 'cf-challenge')]"
    ]

    for indicator in blocking_indicators:
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, indicator)))
            print("⚠ Human verification detected. Clearing cookies and retrying after 1 minute...")

            # Clear cookies and storage
            driver.delete_all_cookies()
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")

            # Wait for 1 minute
            time.sleep(60)

            # Recreate a driver with a fresh session
            driver.quit()
            driver = setup_driver()

            # Retry the request
            driver.get(current_url)
            human_like_interaction(driver)
            random_delay(1, 2)

            return True, driver
        except:
            continue

    return False, driver


def scrape_flight_data(driver, search_params):
    """Scrape flight data with blocking recovery based on user input"""
    departure_airport = search_params['departure_airport']
    arrival_airport = search_params['arrival_airport']
    date_from_str = search_params['date_from']
    date_to_str = search_params['date_to']
    nights = int(search_params['nights'])
    stops = search_params['stops']
    flight_hours = int(search_params['flight_hours'])
    country = search_params.get('country', 'USA')  # Default to USA if not provided
    departure_airport_optional = search_params.get('departure_airport_optional')
    arrival_airport_optional = search_params.get('arrival_airport_optional')

    try:
        departure_date = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        return_date = departure_date + timedelta(days=nights)
        return_date_str = return_date.strftime('%Y-%m-%d')

        stops = search_params['stops']
        stops_param = ""
        if stops:
            stops_list = []
            if '0' in stops:
                stops_list.append("0")
            for stop in stops:
                if stop.isdigit() and stop != '0':
                    stops_list.append(stop)

            if stops_list:
                stops_param = ";stops=" + ",".join(stops_list)

        base_url = "https://www.kayak.com/flights"
        if country == 'Canada':
            base_url = "https://www.ca.kayak.com/flights"

        url = f"{base_url}/{departure_airport}-{arrival_airport}/{date_from_str}/{return_date_str}/2adults?sort=price_a&fs=legdur<={flight_hours * 60}{stops_param};virtualinterline=-virtualinterline;airportchange=-airportchange"
        if country in ['USA', 'Canada'] and departure_airport_optional and arrival_airport_optional:
            url = f"{base_url}/{departure_airport}-{arrival_airport}/{date_from_str}/{departure_airport_optional}-{arrival_airport_optional}/{return_date_str}/2adults?sort=price_a&fs=legdur<={flight_hours * 60}{stops_param};virtualinterline=-virtualinterline;airportchange=-airportchange"

        print(f"🌐 Accessing: {url}")
        driver.get(url)
        random_delay(1, 2)
        human_like_interaction(driver)

        # Check for blocking
        blocked, driver = handle_possible_blocking(driver, url)
        if blocked:
            print("🔄 Retrying after block resolution...")
            random_delay(5, 10)

        # Wait for results
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'nrc6')]")))
        except TimeoutException:
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'prices')]")))
            except:
                print("❌ Timed out waiting for flight results")
                return None, driver

        # Human-like scrolling
        for _ in range(3):
            scroll_amount = random.randint(300, 700)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            random_delay(1, 2)

        flights = driver.find_elements(By.XPATH, "//div[contains(@class, 'nrc6')]")
        if not flights:
            print("❌ No flights found on page")
            return None, driver

        flight = flights[0]
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth'});", flight)
        random_delay(1, 2)

        def safe_extract(xpath, default="Unknown"):
            """Try extracting text from an element, return default if missing"""
            try:
                text = flight.find_element(By.XPATH, xpath).text.strip()
                return text.replace('$', '').replace(',', '').strip()
            except:
                return default

        duration_elements = flight.find_elements(By.XPATH, ".//div[contains(@class, 'xdW8')]/div[contains(@class, 'vmXl')]")
        duration1 = duration_elements[0].text.strip() if len(duration_elements) > 0 else "Unknown"
        duration2 = duration_elements[1].text.strip() if len(duration_elements) > 1 else "Unknown"

        month_name = departure_date.strftime('%B')  # MARK: Get the full month name
        formatted_month = month_name[:3]  # MARK: Take first 3 letters of month
        formatted_date = f"{departure_date.day:02d}-{formatted_month}-{str(departure_date.year)[-2:]}"

        price_xpath = ".//div[contains(@class, 'e2GB-price-text-container')]/div[contains(@class, 'e2GB-price-text')]"
        airline_xpath = ".//div[contains(@class, 'J0g6-operator-text')]"
        if country == 'USA':
            airline_xpath = ".//div[contains(@class, 'c_cgF c_cgF-mod-variant-default')]"

        price_text = safe_extract(price_xpath)
        if country == 'Canada':
            price_text = price_text.replace('C ', '') # Remove 'C ' for Canadian prices
        airline = safe_extract(airline_xpath)

        # Construct the Arrival Airport string for the Excel output
        excel_arrival_airport = arrival_airport
        if departure_airport_optional and arrival_airport_optional:
            excel_arrival_airport = f"{arrival_airport} x {arrival_airport_optional}"

        flight_data = {
            'Date': formatted_date,
            'Departure Airport': departure_airport,
            'Arrival Airport': excel_arrival_airport,
            'Nights': nights,
            'Airline': airline,
            'Price': float(price_text),
            'Departure Time': duration1,
            'Arrival Time': duration2
        }

        print(f"✅ Found flight: {flight_data['Airline']} for {flight_data['Price']}")
        print(f"Found flight data: {flight_data}")
        return flight_data, driver

    except Exception as e:
        print(f"❌ Error scraping for {date_from_str}: {str(e)}")
        return None, driver


@app.route('/', methods=['GET', 'POST'])
def index():
    usa_airports = ["JFK", "EWR", "BOS", "MIA", "MCO", "ORD", "IAH", "IAD", "DEN", "DTW", "PHL", "LAS", "LAX", "SFO", "ATL", "DFW", "SWF"]
    canada_airports = ["YYZ", "YVR", "YOW", "YUL", "YHZ", "YEG", "YYC"]
    selected_country = request.form.get('country')

    if request.method == 'POST':
        departure_airport = request.form['departure_airport']
        arrival_airport = request.form['arrival_airport']
        date_from = request.form['date_from']
        date_to = request.form['date_to']
        nights = int(request.form['nights'])
        stops = request.form.getlist('stops')
        flight_hours = int(request.form['flight_hours'])
        country = request.form.get('country')
        departure_airport_optional = request.form.get('departure_airport_optional')
        arrival_airport_optional = request.form.get('arrival_airport_optional')

        search_params = {
            'departure_airport': departure_airport,
            'arrival_airport': arrival_airport,
            'date_from': date_from,
            'date_to': date_to,
            'nights': nights,
            'stops': stops,
            'flight_hours': flight_hours,
            'country': country,
            'departure_airport_optional': departure_airport_optional,
            'arrival_airport_optional': arrival_airport_optional
        }

        print(f"Form Data: {search_params}")

        all_flights = []
        driver = setup_driver()

        if driver:
            try:
                start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                current_date = start_date

                while current_date <= end_date:
                    delay = random.uniform(1.5, 4.5)
                    print(f"⏳ Waiting {delay:.1f} seconds before next search...")
                    time.sleep(delay)

                    current_search_params = search_params.copy()
                    current_search_params['date_from'] = current_date.strftime('%Y-%m-%d')
                    current_search_params['date_to'] = (current_date + timedelta(days=nights)).strftime('%Y-%m-%d')

                    flight_data, driver = scrape_flight_data(driver, current_search_params)
                    if flight_data:
                        all_flights.append(flight_data)

                    current_date += timedelta(days=1)

                    if random.random() > 0.7:
                        human_like_interaction(driver)

                print(f"Number of flights found: {len(all_flights)}")

                if all_flights:
                    # Save to Excel using openpyxl for formatting
                    df = pd.DataFrame(all_flights)
                    df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%y')
                    output_file = "kayak_flights_results(Testing).xlsx"

                    wb = Workbook()
                    ws = wb.active

                    # Define border style
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
                        ws.append(row)
                        for c_idx, cell in enumerate(ws[r_idx + 1]):
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                            if r_idx == 0:
                                cell.font = Font(bold=True)

                    # Format the Date column to display as DD-MMM-YY
                    date_column = ws['A']
                    for cell in date_column:
                        cell.number_format = 'DD-MMM-YY'

                    wb.save(output_file)
                    print(f"💾 Saved results to {output_file}")
                    return render_template('results.html', output_file=output_file)
                else:
                    return render_template('results.html')

            except Exception as e:
                print(f"❌ An error occurred: {e}")
                return render_template('results.html', error=str(e))
            finally:
                if driver:
                    try:
                        driver.quit()
                    except Exception as e:
                        print(f"Error while closing WebDriver: {e}")
        else:
            return "Failed to setup driver."

    return render_template('index.html', usa_airports=usa_airports, canada_airports=canada_airports, selected_country=selected_country)

@app.route('/download_results')
def download_results():
    output_file = "kayak_flights_results(Testing).xlsx"
    try:
        return send_file(output_file, as_attachment=True, download_name=output_file)
    except FileNotFoundError:
        return "Error: Results file not found."

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
